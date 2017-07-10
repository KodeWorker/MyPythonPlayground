import openpyxl
import datetime

# Getting Sheets from Workbook ================================================
print('*** Sec.1 ***')

wb = openpyxl.load_workbook('example.xlsx')
print('Sheets: %s' %wb.get_sheet_names())

sheet = wb.get_sheet_by_name('Sheet3')
activeSheet = wb.active

print('Selected Sheet: %s' %(sheet))
print('Active Sheet: %s' %(activeSheet))

print('*** Sec.1 Ends Here ***\n')

# Getting Cells from the Sheets ===============================================
print('*** Sec.2 ***')

time = activeSheet['A1'].value
date = datetime.datetime.strptime(time, '%d/%m/%Y %I:%M:%S %p')
print(date)

c = activeSheet['B1']
print('Row ' + str(c.row) + ', Column ' + c.column + ' is ' + c.value)
print('Cell ' + c.coordinate + ' is ' + c.value)

row = 1
col = 2
print('Row: %d, Column: %d -> Value: %s' %(row, col, activeSheet.cell(row=row, column=col).value))

print('--- PRINT SECOND COLUMN IN THE ACTIVE SHEET ---')
for i in range(1, activeSheet.max_row+1):
    print(i, activeSheet.cell(row=i, column=2).value)

print('*** Sec.2 Ends Here ***\n')

# Getting Rows and Columns from the Sheets ====================================

selected_cells = tuple(activeSheet['A1': 'C3'])
print(selected_cells)

print('--- PRINT SELECTED REGION ---')
for rowOfCellObjects in activeSheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('--- END OF ROW ---')